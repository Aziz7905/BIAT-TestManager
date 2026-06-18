from io import BytesIO
from types import SimpleNamespace

from django.test import SimpleTestCase

from apps.specs.services.parsers.csv_parser import CSVSpecificationSourceParser
from apps.specs.services.parsers.xlsx_parser import XLSXSpecificationSourceParser


def _review(record):
    return record.record_metadata["review"]


def _find_record(records, needle):
    return next(record for record in records if needle in record.content)


class TabularSpecificationSourceParserTests(SimpleTestCase):
    def test_csv_parser_extracts_faithful_table_region_without_guessing(self):
        payload = BytesIO(
            (
                "Specification export,,,\n"
                "Reference,Title,Description,Section\n"
                "REQ-001,Select eligible claims,Select claims with pending status,Claims\n"
                "Reference,Title,Description,Section\n"
                "REQ-002,Generate invoices,Generate invoice file,Billing\n"
            ).encode("utf-8")
        )
        payload.name = "sample.csv"
        source = SimpleNamespace(name="Sample CSV", file=payload)

        result = CSVSpecificationSourceParser().parse(source)

        self.assertEqual(result.source_metadata["parser_strategy"], "structural_grid_v1")
        regions = result.column_mapping["regions"]
        self.assertEqual(len(regions), 1)
        self.assertEqual(regions[0]["structural_type"], "table")
        self.assertTrue(regions[0]["needs_mapping"])
        self.assertEqual(
            regions[0]["header_candidates"][0]["values"],
            ["Reference", "Title", "Description", "Section"],
        )

        # Two data rows (the repeated header row is skipped structurally).
        self.assertEqual(len(result.records), 2)
        first = _find_record(result.records, "REQ-001")
        self.assertIn("Reference: REQ-001", first.content)
        self.assertIn("Title: Select eligible claims", first.content)
        # No semantic guessing: the parser does not infer an external reference.
        self.assertEqual(first.external_reference, "")
        self.assertFalse(first.is_selected)
        self.assertTrue(_review(first)["needs_mapping"])
        self.assertFalse(_review(first)["confirmed"])
        self.assertIsNone(_review(first)["record_type"])

    def test_xlsx_parser_detects_table_and_text_regions_faithfully(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"
        sheet.append(["Generated export"])
        sheet.append([])
        sheet.append(["ID", "Feature", "Preconditions", "Steps", "Expected Result"])
        sheet.append(["RX-01", "Prepare settlement", "Claims exist", "Open job screen", "Settlement is generated"])
        sheet.append(["ID", "Feature", "Preconditions", "Steps", "Expected Result"])
        sheet.append(["RX-02", "Review settlement", "", "Open generated file", "Data is visible"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        payload.name = "sample.xlsx"
        source = SimpleNamespace(name="Sample XLSX", file=payload)

        result = XLSXSpecificationSourceParser().parse(source)

        self.assertEqual(result.source_metadata["parser_strategy"], "structural_grid_v1")
        types = {region["structural_type"] for region in result.column_mapping["regions"]}
        self.assertEqual(types, {"text_block", "table"})

        # The single-cell intro row is preserved as a low-friction context record.
        context = _find_record(result.records, "Generated export")
        self.assertTrue(context.is_selected)
        self.assertFalse(_review(context)["needs_mapping"])
        self.assertEqual(_review(context)["record_type"], "context")

        rx01 = _find_record(result.records, "RX-01")
        self.assertIn("ID: RX-01", rx01.content)
        self.assertIn("Feature: Prepare settlement", rx01.content)
        self.assertEqual(rx01.external_reference, "")
        self.assertFalse(rx01.is_selected)
        self.assertTrue(_review(rx01)["needs_mapping"])

    def test_xlsx_parser_skips_hidden_sheets_and_reports_metadata(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Visible requirements"
        sheet.append(["Reference", "Title", "Business Rule", "Expected Result"])
        sheet.append(["REQ-10", "Create transfer", "Customer creates a transfer", "Transfer is saved"])
        hidden = workbook.create_sheet("Hidden draft")
        hidden.sheet_state = "hidden"
        hidden.append(["Reference", "Title"])
        hidden.append(["DRAFT-1", "Should not parse"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        payload.name = "requirements.xlsx"
        source = SimpleNamespace(name="Requirements", file=payload)

        result = XLSXSpecificationSourceParser().parse(source)

        self.assertEqual(result.source_metadata["visible_sheet_count"], 1)
        self.assertEqual(len(result.records), 1)
        record = result.records[0]
        self.assertIn("Reference: REQ-10", record.content)
        self.assertIn("Title: Create transfer", record.content)
        # No business guessing — reference/title are mapped by the user, not inferred.
        self.assertEqual(record.external_reference, "")
        self.assertTrue(_review(record)["needs_mapping"])
