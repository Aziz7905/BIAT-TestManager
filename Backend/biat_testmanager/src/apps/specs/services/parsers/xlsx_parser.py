from .base import ParsedSourceResult, SpecificationSourceParseError, clean_text
from .tabular_structure import (
    GridCell,
    build_records_for_region,
    column_to_letter,
    segment_regions,
    trim_empty_grid,
)


class XLSXSpecificationSourceParser:
    def parse(self, source):
        if not source.file:
            raise SpecificationSourceParseError("An XLSX file is required.")

        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise SpecificationSourceParseError(
                "XLSX parsing requires the openpyxl package."
            ) from error

        workbook = load_workbook(filename=source.file, data_only=True)
        source.file.seek(0)
        formula_workbook = load_workbook(filename=source.file, data_only=False)
        source.file.seek(0)

        records = []
        regions_descriptor: list[dict] = []
        sheet_metadata: list[dict] = []
        visible_sheet_count = 0

        for sheet in workbook.worksheets:
            sheet_info = {"name": sheet.title, "state": sheet.sheet_state, "region_count": 0, "record_count": 0}

            if sheet.sheet_state != "visible":
                sheet_info.update({"skipped": True, "skip_reason": "hidden_sheet"})
                sheet_metadata.append(sheet_info)
                continue

            visible_sheet_count += 1
            formula_sheet = formula_workbook[sheet.title]
            grid, extraction_metadata = _extract_sheet_grid(sheet, formula_sheet)
            sheet_info.update(extraction_metadata)

            grid = trim_empty_grid(grid)
            if not grid:
                sheet_info.update({"skipped": True, "skip_reason": "empty_sheet"})
                sheet_metadata.append(sheet_info)
                continue

            sheet_record_count = 0
            for region in segment_regions(grid, container=sheet.title):
                region_records, region_metadata = build_records_for_region(region)
                records.extend(region_records)
                sheet_record_count += len(region_records)
                regions_descriptor.append(
                    {
                        **region_metadata,
                        "needs_mapping": region_metadata["structural_type"] == "table",
                    }
                )

            sheet_info["region_count"] = sum(
                1 for region in regions_descriptor if region["container"] == sheet.title
            )
            sheet_info["record_count"] = sheet_record_count
            sheet_metadata.append(sheet_info)

        return ParsedSourceResult(
            records=records,
            source_metadata={
                "filename": source.file.name.split("/")[-1],
                "format": "xlsx",
                "sheet_count": len(workbook.worksheets),
                "visible_sheet_count": visible_sheet_count,
                "parser_strategy": "structural_grid_v1",
                "sheets": sheet_metadata,
            },
            column_mapping={"regions": regions_descriptor},
        )


def _extract_sheet_grid(sheet, formula_sheet):
    hidden_rows = {
        row_index
        for row_index, dimension in sheet.row_dimensions.items()
        if getattr(dimension, "hidden", False)
    }
    hidden_columns = {
        column_name
        for column_name, dimension in sheet.column_dimensions.items()
        if getattr(dimension, "hidden", False)
    }
    # Build the set of merged coordinates from range bounds. Iterating cells
    # would yield MergedCell objects (non-top-left cells of a merged range),
    # which do not expose coordinate/column_letter the way regular cells do.
    merged_coordinates = {
        f"{column_to_letter(column)}{row}"
        for merged_range in sheet.merged_cells.ranges
        for row in range(merged_range.min_row, merged_range.max_row + 1)
        for column in range(merged_range.min_col, merged_range.max_col + 1)
    }

    grid: list[list[GridCell]] = []
    hidden_cell_count = 0
    formula_count = 0

    for row in sheet.iter_rows():
        row_cells: list[GridCell] = []
        for cell in row:
            column_letter = column_to_letter(cell.column)
            coordinate = f"{column_letter}{cell.row}"
            hidden_row = cell.row in hidden_rows
            hidden_column = column_letter in hidden_columns
            if hidden_row or hidden_column:
                hidden_cell_count += 1

            formula = ""
            formula_value = formula_sheet[coordinate].value
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formula = formula_value
                formula_count += 1

            value = clean_text(cell.value)
            row_cells.append(
                GridCell(
                    row=cell.row,
                    column=cell.column,
                    coordinate=coordinate,
                    raw_value=value,
                    displayed_value=value,
                    formula=formula,
                    merged=coordinate in merged_coordinates,
                    hidden_row=hidden_row,
                    hidden_column=hidden_column,
                )
            )
        grid.append(row_cells)

    metadata = {
        "hidden_row_count": len(hidden_rows),
        "hidden_column_count": len(hidden_columns),
        "hidden_cell_count": hidden_cell_count,
        "formula_count": formula_count,
        "merged_cell_count": len(merged_coordinates),
        "raw_max_row": sheet.max_row,
        "raw_max_column": sheet.max_column,
    }
    return grid, metadata
